import platform
import PyQt5
import forex_python.converter
from PyQt5 import QtWidgets
from PyQt5 import QtCore, QtGui, QtWidgets
from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject, QObject, QPoint, QRect, QSize, QTime, QUrl, Qt, QEvent)
from PySide2.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient)
from PySide2.QtWidgets import *
import sys, time, json, lxml
from forex_python.converter import CurrencyRates, CurrencyCodes
from forex_python.bitcoin import BtcConverter
import requests
from bs4 import BeautifulSoup
from pyowm import OWM
import pyowm
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from platform import python_version
import openpyxl

from ui_functions import *
from Currencies import *
from ui_weather import *
from ui_splash_screen import Ui_SplashScreen
from ui_main import Ui_MainWindow


counter = 0
jumper = 10
c = CurrencyRates()
b = BtcConverter()
l_time = time.localtime()
StartTime = time.time()
current_time = time.strftime('%H:%M', l_time)
dict_of_changes = {
    'usd': [],
    'gbp': [],
    'eur': [],
    'try': [],
    'jpy': [],
    'aud': [],
    'btc': [],
    'eth': [],
    'oilbrent': [],
    'oilwti': [],
    'gold': [],
    'silver': [],
    'platinum': [],
    'palladium': [],
    'gas': [],
    'sberbank': [],
    'gazprom': [],
    'lukoil': [],
    'yandex': [],
    'tesla': [],
    'apple': [],
    'time': []

}
list_currency = list(dict_of_changes)
timer_up = PyQt5.QtCore.QTimer()
timer_countdown = PyQt5.QtCore.QTimer()
timer_time_now = PyQt5.QtCore.QTimer()







def excel():
    book = openpyxl.Workbook()

    items_time = dict_of_changes[list_currency[len(list_currency)-1]]

    sheet = book.active

    count = 0
    count_number = 2
    while count != len(list_currency) - 1:
        sheet['A{0}'.format(count_number)] = list_currency[count].upper()

        count_number+=1
        count+=1

    count_time = 0
    count_column = 2
    while count_time != len(items_time):
        sheet.cell(row=1,column=count_column).value = items_time[count_time]

        count_column+=1
        count_time+=1


    count_value = 0
    count_row = 2
    count_column_value = 2
    indications_value = dict_of_changes[list_currency[count_value]]
    count_indications_value = 0
    while count_value != len(list_currency)-1:
        if count_indications_value == len(indications_value):
            count_row+=1
            count_column_value = 2
            count_indications_value = 0
            count_value+=1

        indications_value = dict_of_changes[list_currency[count_value]]
        sheet.cell(row=count_row, column=count_column_value).value = indications_value[count_indications_value]
        count_column_value+=1
        count_indications_value+=1

    sheet['A1'].value = 'Значения/Время'
    sheet['B23'].value = ''
    book.save('Currency changes.xlsx')
    book.close()



class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)


        self.url = 'https://ru.investing.com/news/forex-news'
        self.HEADERS = {
            'accept': '*/*',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36'
        }

        self.req = requests.get(self.url, headers=self.HEADERS)
        self.soup = BeautifulSoup(self.req.text, 'lxml')

        self.ui.Btn_Toggle.clicked.connect(lambda: UIFunctions.toggleMenu(self, 150, True))

        self.ui.btn_Home.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.ui.page_Home))
        self.ui.btn_Excel.clicked.connect( excel )
        self.ui.btn_Converter.clicked.connect(lambda: self.ui.stackedWidget.setCurrentWidget(self.ui.page_Converter))

        self.ui.btn_Information.clicked.connect(lambda: UIFunctions.toggleInformation(self, 300, True))

        self.ui.btn_converter.clicked.connect( self.convent_currency )

        self.ui.btn_weather.clicked.connect(lambda: Waether.weather(self))

        self.ui.btn_get_percent.clicked.connect( self.percentage_difference )


        self.update()



        timer_countdown.timeout.connect(self.countdown)
        timer_up.timeout.connect(self.update)
        timer_time_now.timeout.connect(self.time_now)


        def setValue(self, slider, labelPercentage, progressBarName, color):

            value = slider.value()

            sliderValue = int(value)

            htmlText = """<p align="center"><span style=" font-size:50pt;">{VALUE}</span><span style=" font-size:40pt; vertical-align:super;">%</span></p>"""
            labelPercentage.setText(htmlText.replace("{VALUE}", str(sliderValue)))

            self.progressBarValue(sliderValue, progressBarName, color)

    def progressBarValue(self, value, widget, color):

        styleSheet = """
        QFrame{
        	border-radius: 110px;
        	background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:{STOP_1} rgba(85, 170, 255, 255), stop:{STOP_2} {COLOR});
        }
        """

        progress = (100 - value) / 100.0

        # GET NEW VALUES
        stop_1 = str(progress - 0.001)
        stop_2 = str(progress)

        if value == 100:
            stop_1 = "1.000"
            stop_2 = "1.000"

        newStylesheet = styleSheet.replace("{STOP_1}", stop_1).replace("{STOP_2}", stop_2).replace("{COLOR}", color)

        widget.setStyleSheet(newStylesheet)

    def countdown(self):
        global StartTime
        time_countdown = int(time.time() - StartTime)

        minutes = (time_countdown % 3500) // 60
        seconds = time_countdown % 60

        if minutes == 1:
            StartTime = time.time()
            time_countdown = int(time.time() - StartTime)
            minutes = (time_countdown % 3500) // 60
            seconds = time_countdown % 60
            time_string = '0{0}:{1}'.format(minutes, str(59 - seconds))
        else:
            time_string = '0{0}:{1}'.format(minutes, str(59 - seconds))
            if seconds >= 50:
                time_string = '0{0}:0{1}'.format(minutes, str(59 - seconds))

        return self.ui.label_update_inf.setText(time_string)

    def time_now(self):

        time_now = time.localtime()
        true_time_now = time.strftime('%H:%M', time_now)

        return self.ui.label_time_inf.setText(str(true_time_now))

    def percentage_difference(self):
        try:
            self.ui.lineEdit_enter_currency_percent.setStyleSheet("QLineEdit{\n"
"    border: rgb(120, 230, 130);\n"
"    background-color: rgb(35, 35, 35);\n"
"    color: #FFF;\n"
"}\n"
"\n"
"QLineEdit:hover{\n"
"    border: 2px solid rgb(177, 244, 92);\n"
"}\n"
"\n"
"QLineEdit:focus{\n"
"    border: 2px solid rgb(177, 244, 92);    \n"
"    background-color: rgb(55, 55, 55);\n"
"}\n"
"\n"
"\n"
"")
            self.ui.label_percent_inf.setStyleSheet("color: #FFF")
            self.ui.circular_percentProgress.setStyleSheet("QFrame{\n"
"    border-radius: 80%;\n"
"    background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:0.100 rgba(255, 0, 127, 0), stop:0.100 rgba(120, 230, 130, 255));\n"
"}")
            self.ui.btn_get_percent.setStyleSheet("QPushButton {\n"
"    border-radius: 20%;\n"
"    border: 2px solid rgb(120, 230, 130);\n"
"    background-color:rgb(44, 44, 44);\n"
"    color: #FFF\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background-color: rgb(55, 55, 55);\n"
"    border: 2px solid rgb(177, 244, 92);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background-color:  rgb(177, 244, 92)\n"
"}\n"
"\n"
"\n"
"")
            list_of_enter_percent_currency = dict_of_changes[self.ui.lineEdit_enter_currency_percent.text().casefold()]
            initial_number = float(list_of_enter_percent_currency[0])
            last_number = float(list_of_enter_percent_currency[len(list_of_enter_percent_currency)-1])
            if initial_number > last_number:
                percent = str(round((initial_number - last_number) / last_number * 100, 2))
                if percent == '0,00' or percent == '0,0':
                    self.ui.label_percent_inf.setText('STABLE')
                else:
                    self.ui.label_percent_inf.setText('-{0}%'.format(percent))
            elif last_number > initial_number:
                percent = str(round((initial_number - last_number) / initial_number * 100, 2)).replace('-', '')
                if percent == '0,00' or percent == '0,0':
                    self.ui.label_percent_inf.setText('STABLE')
                else:
                    self.ui.label_percent_inf.setText('+{0}%'.format(percent))
            elif initial_number == last_number:
                self.ui.label_percent_inf.setText('STABLE')
        except KeyError:
            self.ui.label_percent_inf.setStyleSheet("color: rgb(215, 50, 75)")
            self.ui.label_percent_inf.setText('0,00%')
            self.ui.lineEdit_enter_currency_percent.setPlaceholderText('Invalid data')
            self.ui.lineEdit_enter_currency_percent.setText('')
            self.ui.lineEdit_enter_currency_percent.setStyleSheet("QLineEdit{\n"
"    border: rgb(215, 50, 75);\n"
"    background-color: rgb(35, 35, 35);\n"
"    color: rgb(215, 50, 75);\n"
"}\n"
"\n"
"QLineEdit:hover{\n"
"    border: 2px solid rgb(215, 50, 75);\n"
"}\n"
"\n"
"QLineEdit:focus{\n"
"    border: 2px solid rgb(215, 50, 75);\n"
"    background-color: rgb(55, 55, 55);\n"
"}\n"
"\n"
"\n"
"")
            self.ui.circular_percentProgress.setStyleSheet("QFrame{\n"
"    border-radius: 80%;\n"
"    background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:0.100 rgba(255, 0, 127, 0), stop:0.100 rgba(215, 50, 75, 255));\n"
"}")
            self.ui.btn_get_percent.setStyleSheet("QPushButton {\n"
"    border-radius: 20%;\n"
"    border: 2px solid rgb(215, 50, 75);\n"
"    background-color:rgb(44, 44, 44);\n"
"    color: rgb(215, 50, 75);\n"
"}\n"
"\n"
"QPushButton:hover {\n"
"    background-color: rgb(55, 55, 55);\n"
"    border: 2px solid rgb(215, 50, 75);\n"
"}\n"
"\n"
"QPushButton:pressed {\n"
"    background-color:  rgb(177, 244, 92)\n"
"}\n"
"\n"
"\n"
"")


    def convent_currency(self):
        global c, b

        try:
            self.ui.label_converter_get_it.setStyleSheet("QLabel {\n"
 "\n"
 "	border: 2px solid rgb(120, 230, 130);\n"
 "	border-radius: 20%;\n"
 "	color: #FFF;\n"
 "	background-color:rgb(44, 44, 44);\n"
 "	padding-left: 15px;\n"
 "}")
            input_currency = self.ui.lineEdit_from_the_currency.text().upper()  # Из какой
            amount = float(self.ui.lineEdit_the_amount.text())  # Сумма
            output_currency = self.ui.lineEdit_to_the_currency.text().upper()  # В какую
            if input_currency == 'BTC' and output_currency == 'ETH':
                output_amount = round((Currencies.currency_difference_BTC(self, b) * amount) / float(self.Currencies.currency_difference_ETH(self, c)[1]), 3)
                self.ui.label_converter_get_it.setText(str(output_amount))
            elif input_currency == 'ETH' and output_currency == 'BTC':
                output_amount = round((float(Currencies.currency_difference_ETH(c)[1]) * amount) / self.Currencies.currency_difference_BTC(self, b), 5)
                self.ui.label_converter_get_it.setText(str(output_amount))
            elif output_currency == 'BTC':
                output_amount = round(b.convert_to_btc(amount, input_currency), 2)
                self.ui.label_converter_get_it.setText(str(output_amount))
            elif input_currency == output_currency and output_currency == input_currency:
                output_amount = amount * 1
                self.ui.label_converter_get_it.setText(str(output_amount))
            elif input_currency == 'BTC':
                output_amount = round(b.convert_btc_to_cur(amount, output_currency), 2)
                self.ui.label_converter_get_it.setText(str(output_amount))
            elif input_currency == 'ETH':
                convert = c.get_rate('USD', output_currency)
                output_amount = round(float(Currencies.currency_difference_ETH(self, c)[1]) * amount * float(convert), 2)
                self.ui.label_converter_get_it.setText(str(output_amount))
            elif output_currency == 'ETH':
                convert = c.get_rate(input_currency, 'USD')
                output_amount = round(float(Currencies.currency_difference_ETH(self, c)[1]) / (float(convert) * float(amount)), 2)
                self.ui.label_converter_get_it.setText(str(output_amount))
            else:
                convert = c.get_rate(input_currency, output_currency)
                output_amount = round(convert * amount, 2)
                self.ui.label_converter_get_it.setText(str(output_amount))

        except ValueError:
            self.ui.label_converter_get_it.setStyleSheet("QLabel {\n"
"\n"
"	border: 2px solid rgb(215, 50, 75);\n"
"	border-radius: 20%;\n"
"	color: rgb(215, 50, 75);\n"
"	background-color:rgb(44, 44, 44);\n"
"	padding-left: 15px;\n"
"}")
            self.ui.label_converter_get_it.setText('The amount must be a digit')
        except forex_python.converter.RatesNotAvailableError:
            self.ui.label_converter_get_it.setStyleSheet("QLabel {\n"
"\n"
"	border: 2px solid rgb(215, 50, 75);\n"
"	border-radius: 20%;\n"
"	color: rgb(215, 50, 75);\n"
"	background-color:rgb(44, 44, 44);\n"
"	padding-left: 0px;\n"
"   font-size: 13px;\n"                                                      
"}")
            self.ui.label_converter_get_it.setText('There is no such currency or it is typed incorrectly')




    def update(self):
        global c, b, l_time, current_time, dict_of_changes, Currencies

        c = CurrencyRates()
        b = BtcConverter()
        l_time = time.localtime()
        current_time = time.strftime('%H:%M', l_time)

        self.ui.C_USD.setText(str(Currencies.currency_difference_USD(self, c)) + '  ₽')
        self.ui.C_GBP.setText(str(Currencies.currency_difference_GBP(self, c)) + '  ₽')
        self.ui.C_EUR.setText(str(Currencies.currency_difference_EUR(self, c)) + '  ₽')
        self.ui.C_TRY.setText(str(Currencies.currency_difference_TRY(self, c)) + '  ₽')
        self.ui.C_JPY.setText(str(Currencies.currency_difference_JPY(self, c)) + '  ₽')
        self.ui.C_AUD.setText(str(Currencies.currency_difference_AUD(self, c)) + '  ₽')
        self.ui.C_BTC.setText(str(Currencies.currency_difference_BTC(self, b)) + '  ₽')
        self.ui.C_ETH.setText(str(Currencies.currency_difference_ETH(self, c)[0]) + '  ₽')
        self.ui.C_Gold.setText(str(Currencies.Gold(self)) + '  ₽')
        self.ui.C_Silver.setText(str(Currencies.Silver(self)) + '  ₽')
        self.ui.C_Platinum.setText(str(Currencies.Platinum(self)) + '  ₽')
        self.ui.C_Palladium.setText(str(Currencies.Palladium(self)) + '  ₽')
        self.ui.C_OilBrent.setText(str(Currencies.OilBrent(self)) + '  ₽')
        self.ui.C_OilWTI.setText(str(Currencies.OilWTI(self)) + '  ₽')
        self.ui.C_Gas.setText(str(Currencies.Gas(self)) + '  ₽')
        self.ui.C_Sberbank.setText(str(Currencies.Sberbank(self)) + '  ₽')
        self.ui.C_Gazprom.setText(str(Currencies.Gazprom(self)) + '  ₽')
        self.ui.C_Lukoil.setText(str(Currencies.Lukoil(self)) + '  ₽')
        self.ui.C_Yandex.setText(str(Currencies.Yandex(self)) + '  ₽')
        self.ui.C_Tesla.setText(str(Currencies.Tesla(self)) + '  ₽')
        self.ui.C_Apple.setText(str(Currencies.Apple(self)) + '  ₽')

        dict_of_changes['usd'].append(Currencies.currency_difference_USD(self,c))
        dict_of_changes['gbp'].append(Currencies.currency_difference_GBP(self,c))
        dict_of_changes['eur'].append(Currencies.currency_difference_EUR(self,c))
        dict_of_changes['try'].append(Currencies.currency_difference_TRY(self,c))
        dict_of_changes['jpy'].append(Currencies.currency_difference_JPY(self,c))
        dict_of_changes['aud'].append(Currencies.currency_difference_AUD(self,c))
        dict_of_changes['btc'].append(Currencies.currency_difference_BTC(self,b))
        dict_of_changes['eth'].append(Currencies.currency_difference_ETH(self,c)[0])
        dict_of_changes['gold'].append(Currencies.Gold(self))
        dict_of_changes['oilbrent'].append(Currencies.OilBrent(self))
        dict_of_changes['oilwti'].append(Currencies.OilWTI(self))
        dict_of_changes['silver'].append(Currencies.Silver(self))
        dict_of_changes['platinum'].append(Currencies.Platinum(self))
        dict_of_changes['palladium'].append(Currencies.Palladium(self))
        dict_of_changes['gas'].append(Currencies.Gas(self))
        dict_of_changes['sberbank'].append(Currencies.Sberbank(self))
        dict_of_changes['gazprom'].append(Currencies.Gazprom(self))
        dict_of_changes['lukoil'].append(Currencies.Lukoil(self))
        dict_of_changes['yandex'].append(Currencies.Yandex(self))
        dict_of_changes['tesla'].append(Currencies.Tesla(self))
        dict_of_changes['apple'].append(Currencies.Apple(self))
        dict_of_changes['time'].append(current_time)


class SplashScreen(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_SplashScreen()
        self.ui.setupUi(self)

        self.progressBarValue(0)

        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 120))
        self.ui.circularBg.setGraphicsEffect(self.shadow)

        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)

        self.timer.start(15)

        self.show()

    def progress (self):
        global counter
        global jumper
        value = counter

        htmlText = """<p><span style=" font-size:68pt;">{VALUE}</span><span style=" font-size:58pt; vertical-align:super;">%</span></p>"""

        newHtml = htmlText.replace("{VALUE}", str(jumper))

        if(value > jumper):

            self.ui.labelPercentage.setText(newHtml)
            jumper += 10


        if value >= 100: value = 1.000
        self.progressBarValue(value)

        if counter > 100:

            self.timer.stop()
            self.main = MainWindow()

            self.main.show()
            self.close()


        counter += 0.5

    def progressBarValue(self, value):

        styleSheet = """
        QFrame{
        	border-radius: 150px;
        	background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:{STOP_1} rgba(255, 0, 127, 0), stop:{STOP_2} rgba(120, 230, 130, 255));
        }
        """

        progress = (100 - value) / 100.0

        stop_1 = str(progress - 0.001)
        stop_2 = str(progress)

        newStylesheet = styleSheet.replace("{STOP_1}", stop_1).replace("{STOP_2}", stop_2)

        self.ui.circularProgress.setStyleSheet(newStylesheet)



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SplashScreen()
    timer_up.start(60000)
    timer_countdown.start(1000)
    timer_time_now.start(1000)
    sys.exit(app.exec_())


